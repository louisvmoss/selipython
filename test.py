import openpyxl
import time
import random
from http.server import executable
from lib2to3.pgen2 import driver
from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementNotVisibleException
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# driver = webdriver.Chrome(executable_path='C:\\Users\\Administrator\\Desktop\\selenium wit py\\chromedriver.exe')
def pass_data(url,ip,num):
    driver = webdriver.Chrome(executable_path='C:\\Users\\Lenovo\Desktop\\seleniumpython\\chromedriver.exe')
    options = webdriver.ChromeOptions()
    options.add_argument('ignore-certificate-errors')

    driver.header_overrides = {
	'USER-AGENT':'Mozilla/5.0 (iPhone; CPU iPhone OS 9_2 like Mac OS X) AppleWebKit/601.1 (KHTML, like Gecko) CriOS/47.0.2526.70 Mobile/13C71 Safari/601.1.46',
 	'X-FORWARDED-FOR':str("181.176.111.83"),
 	'X-MSISDN':str(num),
 	'Access-Control-Allow-Credentials':'True',
 	'Access-Control-Allow-Headers':'Content-Type, Access-Control-Allow-Headers, Authorization, X-Requested-With',
 	'Access-Control-Allow-Methods':'POST, GET, OPTIONS, DELETE, PUT',
 	'Access-Control-Max-Age':'3600',
	'Cache-Control':'no-cache, must-revalidate, private',
 	'X-FRAME-OPTIONS':'SAMEORIGIN'
	}
    time.sleep(5)
    driver.get(url)
    
    try:
        driver.set_page_load_timeout(60)
        driver.get(url)
        time.sleep(1)
        # myElem = driver.find_element_by_xpath("\\*[@id="main"]\div[2]\div\div[1]\a[1]")
        #myElem.click()
        #time.sleep(6)
        #driver.find_element_by_id("55527").click()
        #driver.find_element_by_xpath("//*[@type='button']").click()
        # driver.find_element_by_id("send_msisdn").click()
        driver.find_element_by_xpath('//*[@id"=""send_msisdn"]').click()  
        # driver.find_element_by_class_name("send_msisdn").click()
        driver.implicitly_wait(10)
        #driver.set_window_size(480, 620)
        #driver.implicitly_wait(20)
        #driver.find_element_by_class_name("sub-btn").click()
        #driver.find_element_by_class_name("btn").click()
        timeout = 8
    except NoSuchElementException as exception:
        print("Button not found and test failed")
    except ElementNotVisibleException as exception:
        print("element not interactable")
    except TimeoutException as exception:
        print("Already clicked")
    except WebDriverException as exception:
        print("Service chromedriver unexpectedly exited")
    driver.close()
ip_range = ["179.49.161.14","156.200.116.68","27.208.89.157"]

def header_val():
    # urls = input("Enter your URL :")
    urls= 'https://1263eec77132.tc-offerz.net/?p=4396&wid=134151&wid_hmac=45f2b37995e7174c2bb87c3a14ebecdb'
    wb = openpyxl.load_workbook("C:/Users/Lenovo/Documents/Book2.xlsx")
    sheet = wb.active
    m_row = sheet.max_row
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=m_row, max_col=1):
        for cell in row:
            secure_random = random.SystemRandom()
            rand_ip = secure_random.choice(ip_range)
            print(rand_ip,"and",cell.value)
        pass_data(urls,rand_ip,cell.value)
header_val()

